import openpyxl, os

from openpyxl import Workbook

os.chdir("D:\\Tutorial\\tf-files\\python\\documents")


workbook = openpyxl.load_workbook('example.xlsx')
sheet1 = workbook.get_sheet_by_name('Sheet1')
# cell = sheet1['A1']
# print(cell.value)
# test = sheet1.cell(row=1, column=2)
for x in range(1,8):
    a = sheet1.cell(row=x, column=1).value
    b = sheet1.cell(row=x, column=2).value
    c = sheet1.cell(row=x, column=3).value
    print(a,b,c)