import openpyxl, os
os.chdir("D:\\Tutorial\\tf-files\\python\\documents")
workbook = openpyxl.Workbook()
workbook.sheetnames
sheet = workbook['Sheet']
sheet['A1'].value = "skruffer"
sheet['B1'].value = "69"
newsheet = workbook.create_sheet()
newsheet.title = 'new shit'
workbook.save('modify.xlsx')
print(workbook.get_sheet_names())
